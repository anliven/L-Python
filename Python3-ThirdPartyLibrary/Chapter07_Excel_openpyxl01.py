# coding=utf-8
from openpyxl import Workbook
import datetime
import os

wb = Workbook()  # 将默认生成一个名为“sheet”的Worksheet
ws = wb.active  # grab the active worksheet

ws['A1'] = 42  # Data can be assigned directly to cells
ws.append([1, 2, 3])  # Rows can also be appended
ws['A2'] = datetime.datetime.now()  # Python types will automatically be converted

wb.save("sample.xlsx")  # Save the file
wb.close()

os.remove("sample.xlsx")

# ### openpyxl
# - openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
# - Homepage: http://www.python-excel.org/
# - Documentation: https://openpyxl.readthedocs.org
#
# ### 一些操作
# 不仅可以读写和创建表格文件，而且可以完成更丰富和复杂的操作：
#   - 支持数字格式和公式；
#   - 支持合并单元格、插入图片、折叠大纲、设置模板等；
#   - 可以创建图表（面积图、直方图、散点图、线图、饼图、雷达图、坐标图、3D图等）、更改图表布局、添加样式等；
#   - 可以与Pandas和NumPy协作，例如将“Pandas dataframe”转换成工作表；
