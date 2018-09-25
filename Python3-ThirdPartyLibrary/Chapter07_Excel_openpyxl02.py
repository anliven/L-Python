# coding=utf-8
import openpyxl

wb = openpyxl.load_workbook(filename="Chapter07_Excel_openpyxl_Test.xlsx")  # 使用load_workbook函数读取已存在的文档
print("# Read-only:", wb.read_only, "# Encode:", wb.encoding)
print("# Worksheet:", wb.worksheets, "# SheetNames:", wb.sheetnames)
print("# Properties:", wb.properties)  # 文档属性

ws = wb.active  # 获取当前活跃的Worksheet对象
print("# current active Worksheet: ", ws)
ws['A1'] = 123456789

ws2 = wb['Sheet-Two']  # 通过Worksheet['表名']获取Worksheet对象
ws2['B1'] = "abc abc abc"

wb.create_sheet("Sheet-3th")  # 创建一个空的表格

wb.save(filename="Chapter07_Excel_openpyxl_Test.xlsx")
wb.close()

# ### 相关的类
# openpyxl包含三个不同层次的类，每个类都包含了对应的属性和方法：
#   - WorkBook：工作簿类，一个Workbook对象代表一个Excel文档；
#   - Worksheet： 表格类
#   - Cell：单元格类
#
# ### Workbook对象
# 操作Excel之前必须先创建一个Workbook对象
#   - 创建一个新Excel文档，直接调用Workbook类，返回一个Workbook对象；
#   - 读取已存在的文档，使用load_workbook函数， 返回一个Workbook对象；
# 常见的Workbook属性和方法
#   - read_only：判断是否以read_only模式打开Excel文档
#   - encoding：获取文档的字符集编码
#   - worksheets：以列表的形式返回所有的Worksheet
#   - sheetnames：以列表的形式返回工作簿中所有的表名
#   - properties：获取文档的元数据（标题，创建者，创建日期等）
#   - active：获取当前活跃的Worksheet对象
#   - get_sheet_names：获取所有表格的名称(建议通过Workbook的sheetnames属性即可获取)
#   - get_sheet_by_name：通过表格名称获取Worksheet对象(建议通过Worksheet['表名']获取)
#   - get_active_sheet：获取活跃的表格(建议通过active属性获取)
#   - remove_sheet：删除一个表格
#   - create_sheet：创建一个空的表格
#   - copy_worksheet：在Workbook内拷贝表格
