# coding=utf-8
import openpyxl

wb = openpyxl.load_workbook(filename="Chapter07_Excel_openpyxl_Test.xlsx")

ws = wb['Sheet-One']
print("# Current active Worksheet: ", ws, "# Title: ", ws.title)
print("Size:{} MaxMinRow:{} MaxMinColumn:{}".format(
    ws.dimensions,
    (ws.max_row, ws.min_row),
    (ws.max_column, ws.min_column)))

for row in ws.rows:  # 按行遍历获取单元格
    print("# rows:", row)
    for cell in list(row):  # 逐个打印单元格信息
        print("## RowsValue:{0} RowColumn:{1} Coordinate:{2} Parent:{3}".format(
            cell.value,
            (cell.row, cell.column),
            cell.coordinate,
            cell.parent))

for row in ws.values:
    print("# rows value:", *row)

for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=4):
    print("# ws.iter_rows: ", list(row))

ws2 = wb['Sheet-Two']
print("# ws2 Single-Cell: ", ws2['A1'])  # 通过坐标获取单个Cell对象
print("# ws2 Single-Cell: ", ws2.cell(row=1, column=2))  # 通过cell()获取单个Cell对象
print("# ws2 Multi-Cells: ", ws2['A1:C1'])  # 通过坐标获取多个Cell对象, 以tuple形式返回多行数据

ws3 = wb['Sheet-3th']
ws3['A1'] = "This is a test!"  # 通过坐标设置cell的值
ws3['B1'].value = "aaa bbb ccc"  # 通过value属性设置cell的值
ws3.cell(row=1, column=3, value="1234567890")  # 通过cell()设置cell的值
print("# ws3 Cells: ", [c.value for c in ws3['A1:C1'][0]])

wb.save(filename="Chapter07_Excel_openpyxl_Test.xlsx")
wb.close()

# ### Worksheet对象
# 通过Workbook对象可以获取表格的属性，得到单元格中的数据，修改表格中的内容；
# 常见的Worksheet属性和方法（大部分都是返回Cell对象，一个Cell对象代表一个单元格）
#   - title：WorkSheet的名称
#   - dimensions：表格的大小（含有数据的表格的左上角坐标和右下角坐标）
#   - max_row：表格的最大行
#   - min_row：表格的最小行，起始为1
#   - max_column：表格的最大列
#   - min_column：表格的最小列，起始为1
#   - rows：返回所有有效数据行，有数据时类型为generator，无数据时为tuple
#   - columns：返回所有有效数据列,有数据时类型为generator,无数据时为tuple
#   - values：按行返回所有单元格的内容，类型为tuple
#   - freeze_panes：冻结窗格，主要用于在表格较大时冻结顶部的行或左边的行
#   - iter_rows：按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
#   - iter_columns：按列获取所有的单元格
#   - append：在表格末尾添加数据
#   - merged_cells：合并多个单元格
#   - unmerged_cells：移除合并的单元格
#
# ### Cell对象
# 一个Cell对象代表一个单元格；
# 获取Cell对象：
#   - 使用Excel坐标的方式
#   - 使用Worksheet的cell方法
# 常见的Cell属性和方法
#   - column：单元格所在列，起始为1
#   - row：单元格所在行，起始为1
#   - coordinate：单元格所在坐标，如'A1'
#   - parent：单元格所属的WorkSheet
#   - value：单元格的值
